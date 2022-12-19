import openpyxl
import time
import os

def OpenPersonalDataBase():
    os.system('Personal.xlsx')

def OpenMonitoringWindow():
    os.system('StatusTable.xlsx')

def CheckPerson(personal_Base, result):
    wb = openpyxl.reader.excel.load_workbook(filename=personal_Base)
    sheet = wb.active
    base = []
    n = 2
    while True:
        data = [(sheet['A' + str(n)].value)] + [(sheet['B' + str(n)].value)]
        if data[0] == None:
            break
        base += data
        n += 1
    for c in range(len(base) - 1):
        if str(result) == str(base[c]):
            name = base[c+1]
            return name


def MakeFirstEntry(monitoring_Base, result, name):
    start_time = time.strftime("%d/%m/%Y, %H:%M:%S", time.localtime())
    book = openpyxl.reader.excel.load_workbook(filename=monitoring_Base)
    sheet = book.active
    sheet.insert_rows(2)
    sheet['A2'].value = result
    sheet['B2'].value = name
    sheet['C2'].value = start_time
    sheet['E2'].value = "На территории"
    book.save(monitoring_Base)

def MakeSecondEntry(monitoring_Base, result):
    end_time = time.strftime("%d/%m/%Y, %H:%M:%S", time.localtime())
    book = openpyxl.reader.excel.load_workbook(filename=monitoring_Base)
    sheet = book.active
    persons = []
    n = 2
    while True:
        data = [(sheet['A' + str(n)].value)]
        if data[0] == None:
            break
        persons += data
        n += 1
    for c in range(len(persons)):
        if result == persons[c]:
            sheet['D' + str(c + 2)].value = end_time
            sheet['E'  + str(c + 2)].value = "Вне территории"
            book.save(monitoring_Base)


#print(CheckPerson("Personal.xlsx", 77729183))
#MakeFirstEntry("StatusTable.xlsx", result, CheckPerson("Personal.xlsx", result))
#MakeSecondEntry("StatusTable.xlsx", result)